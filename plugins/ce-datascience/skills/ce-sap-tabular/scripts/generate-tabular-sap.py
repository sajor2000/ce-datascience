#!/usr/bin/env python3
"""Render SAP-tables CSVs into a single multi-sheet .xlsx workbook.

Usage:
    python3 generate-tabular-sap.py <slug> <input_dir> <output_xlsx>

Where <input_dir> contains the three core biostatistics SAP sheets:
01-overview.csv, 02-outputs.csv, and 03-variables.csv. Optional synthetic
sample sheets, 04-file1-long-sample.csv and 05-file2-wide-sample.csv, are
included when present.

If openpyxl is not installed, the script exits 0 with a notice; the CSVs alone
are still a valid output.
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

CORE_SHEET_FILES = [
    ("Overview", "01-overview.csv"),
    ("Outputs", "02-outputs.csv"),
    ("Master Variables", "03-variables.csv"),
]

OPTIONAL_SHEET_FILES = [
    ("File 1 (Long)", "04-file1-long-sample.csv"),
    ("File 2 (Wide)", "05-file2-wide-sample.csv"),
]

SECTION_FILL = {
    "DIAGNOSTIC OUTPUTS": "BDD7EE",
    "SETUP / DIAGNOSTICS": "BDD7EE",
    "TABLE OUTPUTS": "FFE699",
    "MODEL OUTPUTS": "C6E0B4",
    "FIGURE DATA OUTPUTS": "F4B084",
    "SUBGROUP / SENSITIVITY OUTPUTS": "D9EAD3",
}

DEFAULT_SECTION_FILL = "D9E1F2"


def read_csv(path: Path, *, required: bool) -> list[list[str]]:
    if not path.exists():
        if required:
            raise FileNotFoundError(f"Required SAP table missing: {path}")
        return []
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.reader(f))


def is_banner_row(row: list[str], expected_cols: int) -> bool:
    """Workbook-style output section row: first cell has text, the rest blank."""
    if not row or not row[0].strip():
        return False
    padded = row + [""] * max(0, expected_cols - len(row))
    return all(not value.strip() for value in padded[1:expected_cols])


def section_fill(section: str) -> str:
    upper = section.strip().upper()
    if upper in SECTION_FILL:
        return SECTION_FILL[upper]
    for prefix, fill in SECTION_FILL.items():
        if upper.startswith(prefix):
            return fill
    return DEFAULT_SECTION_FILL


def main(argv: list[str]) -> int:
    if len(argv) < 4:
        print("usage: generate-tabular-sap.py <slug> <input_dir> <output_xlsx>", file=sys.stderr)
        return 2

    slug = argv[1]
    input_dir = Path(argv[2])
    output_xlsx = Path(argv[3])

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as e:
        print(
            f"openpyxl not installed ({e}); skipping .xlsx generation. "
            "Run `python3 -m pip install --upgrade openpyxl` to enable.",
            file=sys.stderr,
        )
        return 0

    wb = Workbook()
    # Remove the default sheet
    default = wb.active
    wb.remove(default)

    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)

    sheet_files = CORE_SHEET_FILES + [
        sheet for sheet in OPTIONAL_SHEET_FILES if (input_dir / sheet[1]).exists()
    ]

    for sheet_name, filename in sheet_files:
        ws = wb.create_sheet(sheet_name)
        rows = read_csv(input_dir / filename, required=(sheet_name in {"Overview", "Outputs", "Master Variables"}))

        # Title row
        ws.cell(row=1, column=1, value=f"{sheet_name} -- {slug}")
        ws.cell(row=1, column=1).font = title_font

        if not rows:
            continue

        # Determine if 02-outputs has a legacy 'section' column. The preferred
        # biostatistics workbook format uses section banner rows instead.
        has_section_col = (
            sheet_name == "Outputs"
            and rows
            and rows[0]
            and rows[0][0].strip().lower() == "section"
        )

        # Write header row at row 3
        header = rows[0]
        for col_idx, val in enumerate(header, start=1):
            c = ws.cell(row=3, column=col_idx, value=val)
            c.font = header_font
            c.alignment = Alignment(wrap_text=True, vertical="top")

        # Write data rows starting at row 4
        current_section = None
        out_row = 4
        for raw_row in rows[1:]:
            if sheet_name == "Outputs" and not has_section_col and is_banner_row(raw_row, len(header)):
                section = raw_row[0].strip()
                banner = ws.cell(row=out_row, column=1, value=section)
                banner.font = Font(bold=True)
                fill_color = section_fill(section)
                for col_idx in range(1, len(header) + 1):
                    cell = ws.cell(row=out_row, column=col_idx)
                    cell.fill = PatternFill(
                        start_color=fill_color,
                        end_color=fill_color,
                        fill_type="solid",
                    )
                ws.merge_cells(
                    start_row=out_row,
                    start_column=1,
                    end_row=out_row,
                    end_column=len(header),
                )
                out_row += 1
                continue

            if has_section_col and raw_row and raw_row[0].strip():
                section = raw_row[0].strip()
                if section != current_section:
                    # Insert a banner row with the section name spanning all columns
                    banner = ws.cell(row=out_row, column=1, value=section)
                    banner.font = Font(bold=True)
                    fill_color = section_fill(section)
                    for col_idx in range(1, len(header) + 1):
                        cell = ws.cell(row=out_row, column=col_idx)
                        cell.fill = PatternFill(
                            start_color=fill_color,
                            end_color=fill_color,
                            fill_type="solid",
                        )
                    ws.merge_cells(
                        start_row=out_row,
                        start_column=1,
                        end_row=out_row,
                        end_column=len(header),
                    )
                    out_row += 1
                    current_section = section

            for col_idx, val in enumerate(raw_row, start=1):
                cell = ws.cell(row=out_row, column=col_idx, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            out_row += 1

        # Freeze header
        ws.freeze_panes = ws["A4"]

        # Set column widths
        widths = {1: 18, 2: 16, 3: 24, 4: 24, 5: 50, 6: 40, 7: 50}
        for col_idx, width in widths.items():
            if col_idx <= len(header):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    print(f"Wrote {output_xlsx}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
